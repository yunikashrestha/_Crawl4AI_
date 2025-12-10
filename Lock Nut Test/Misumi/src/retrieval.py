import os
from openpyxl import Workbook,load_workbook
from fastembed import TextEmbedding
from qdrant_client import QdrantClient
client=QdrantClient(url="http://localhost:6333")
collection_name="Misumi Test"

dense_encoder=TextEmbedding("sentence-transformers/all-MiniLM-L6-v2")

def embed_query(query):
    dense_embed=next(dense_encoder.query_embed(query))
    return dense_embed

def retrieval(query):
    dense_vector_query=embed_query(query)
    
    search_results=client.query_points(
        collection_name=collection_name,
        query=dense_vector_query,
        using="dense",
        limit=5,
        with_payload=True
    )
    print(search_results)
    return search_results.points

def save_log(query,results):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_path = os.path.join(script_dir, "..", "test logs", "retrieval_logs.txt")
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    
    with open(log_path,"a",encoding="utf-8") as f:
        f.write(f"User Query:{query}\n")
        f.write("-" * 20 + "\n")
        for rank,result in  enumerate(results,start=1):
            score=result.score
            payload=result.payload      

            part_number=payload.get("part_number","N/A")
            content=payload.get("chunk","N/A")
            url=payload.get("URL","N/A")
            
            f.write(f" Rank: {rank} Score: {score: .4f}\n")
            f.write(f" Content: {content}\n")
            f.write(f"URL: {url}\n")
        f.write(f"{'='*50}\n\n")    

def  save_log_excel(query,results):
    script_dir=os.path.dirname(os.path.abspath(__file__))
    log_path=os.path.join(script_dir,"..","test logs","retrieval_logs_misumi.xlsx")
    os.makedirs(os.path.dirname(log_path),exist_ok=True)

    rows_to_add=[]
    for rank,result in enumerate(results,start=1):
            payload=result.payload

            row_data=[
                query,
                rank,
                round(result.score,4),
                payload.get("part_number","N/A"),
                payload.get("chunk","N/A"),
                payload.get("URL","N/A")
            ]
            rows_to_add.append(row_data)
    headers=["User Query","Rank","Score","Part Number","Retrieved Content","URL"]
    if not os.path.exists(log_path):
         wb=Workbook()
         ws=wb.active
         ws.title="Retrieval Logs"
         ws.append(headers)
         for row in rows_to_add:
              ws.append(row)  
         wb.save(log_path)
    else:
        try:
              wb=load_workbook(log_path)
              ws=wb.active
              for row in rows_to_add:
                   ws.append(row)
              wb.save(log_path)
        except PermissionError:
             print("Close the excel file and try again")


if __name__=="__main__":
    user_query="AWL24"
    results=retrieval(user_query)
    save_log_excel(user_query,results)





